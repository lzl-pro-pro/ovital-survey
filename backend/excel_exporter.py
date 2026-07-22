"""Excel 导出引擎 - 统计汇总表和单张原始调查记录表"""

import os
import io
import json
import logging
import zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from backend.data_manager import (
    get_project, get_survey_points, get_point_detail,
    get_project_stats, get_project_photos, get_default_template
)
from config import EXPORT_FOLDER, BASE_DIR

logger = logging.getLogger(__name__)


# ============ 样式常量 ============

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="微软雅黑", size=16, bold=True)
SUBTITLE_FONT = Font(name="微软雅黑", size=12, bold=True)
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
LABEL_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")


NO_FILL = PatternFill(fill_type=None)


def _apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    """应用单元格样式"""
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def _auto_width(ws, min_width=8, max_width=40):
    """自动调整列宽"""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                # 中文字符算2个宽度
                text = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in text)
                max_len = max(max_len, length)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _safe_xl_image(photo_path):
    """安全加载图片为 openpyxl Image 对象，兼容各种格式"""
    try:
        # 直接尝试 openpyxl 加载
        return XLImage(photo_path)
    except Exception as e1:
        try:
            # PIL 兜底：打开 → 转RGB → 存为PNG到内存 → 再用 openpyxl 加载
            pil_img = PILImage.open(photo_path)
            if pil_img.mode in ("RGBA", "P", "LA", "PA"):
                pil_img = pil_img.convert("RGBA")
            else:
                pil_img = pil_img.convert("RGB")
            buf = io.BytesIO()
            pil_img.save(buf, format="PNG")
            buf.seek(0)
            return XLImage(buf)
        except Exception as e2:
            logger.warning(f"照片加载失败: {photo_path} | 原始错误: {e1} | PIL兜底: {e2}")
            return None


def export_summary(project_id, point_ids=None):
    """
    生成统计汇总表

    Columns: 序号 | 编号 | 位置描述 | 经度 | 纬度 | 高程 | 调查人 | 调查日期 | 状态 | 照片数 | 备注
    """
    project = get_project(project_id)
    if not project:
        raise ValueError("项目不存在")

    # 获取所有调查点
    result = get_survey_points(project_id, per_page=10000)
    points = result["items"]

    # 如果指定了point_ids，过滤
    if point_ids:
        pid_set = set(point_ids)
        points = [p for p in points if p["id"] in pid_set]

    wb = Workbook()
    ws = wb.active
    ws.title = "统计汇总表"

    # 标题行
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"{project['name']} - 野外调查统计汇总表"
    _apply_cell_style(title_cell, font=Font(name="微软雅黑", size=14, bold=True),
                      alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 30

    # 副标题（日期范围）
    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    sub_cell.value = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    _apply_cell_style(sub_cell, font=Font(name="微软雅黑", size=9, color="666666"),
                      alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 22

    # 表头 (第4行)
    headers = ["序号", "编号", "位置描述", "经度", "纬度", "高程(m)",
               "调查人", "调查日期", "状态", "照片数", "备注"]
    header_row = 4

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        _apply_cell_style(cell, HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT, THIN_BORDER)
    ws.row_dimensions[header_row].height = 25

    # 状态映射
    status_map = {
        "pending": "待调查", "in_progress": "进行中",
        "surveyed": "已完成", "skipped": "已跳过"
    }

    # 数据行
    data_start = header_row + 1
    for i, point in enumerate(points):
        row = data_start + i
        detail = get_point_detail(point["id"]) or {}
        records = {r["field_key"]: r["field_value"] for r in detail.get("records", [])}
        photo_count = point.get("photo_count", 0)

        row_data = [
            i + 1,
            point["point_number"],
            records.get("location_desc", ""),
            point["longitude"] if point["longitude"] else "",
            point["latitude"] if point["latitude"] else "",
            point["altitude"] if point["altitude"] else "",
            records.get("investigator", ""),
            records.get("survey_date", "") or (
                point["surveyed_at"][:10] if point.get("surveyed_at") else ""
            ),
            status_map.get(point["status"], point["status"]),
            photo_count,
            records.get("remarks", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            # 交替行颜色
            fill = ALT_ROW_FILL if i % 2 == 1 else NO_FILL
            align = CENTER_ALIGN if col_idx in (1, 4, 5, 6, 9, 10) else LEFT_ALIGN
            _apply_cell_style(cell, NORMAL_FONT, fill, align, THIN_BORDER)

        # GPS列格式
        for c in (4, 5):
            cell = ws.cell(row=row, column=c)
            if cell.value:
                cell.number_format = "0.000000"

        ws.row_dimensions[row].height = 22

    # 统计汇总行
    stats = get_project_stats(project_id)
    summary_start = data_start + len(points) + 1

    summary_rows = [
        ("合计", stats["total"]),
        ("已完成", stats["surveyed"]),
        ("进行中", stats["in_progress"]),
        ("待调查", stats["pending"]),
        ("已跳过", stats["skipped"]),
        ("完成率", f"{stats['surveyed']/max(stats['total'],1)*100:.1f}%"),
    ]

    for j, (label, value) in enumerate(summary_rows):
        row = summary_start + j
        cell_label = ws.cell(row=row, column=1, value=label)
        _apply_cell_style(cell_label, BOLD_FONT, SUMMARY_FILL, CENTER_ALIGN, THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        cell_val = ws.cell(row=row, column=3, value=value)
        _apply_cell_style(cell_val, BOLD_FONT, SUMMARY_FILL, CENTER_ALIGN, THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=11)

        # 填充合并区域样式
        for c in range(2, 12):
            ws.cell(row=row, column=c).fill = SUMMARY_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER

        ws.row_dimensions[row].height = 22

    # 调整列宽
    col_widths = [6, 20, 25, 12, 12, 10, 10, 12, 8, 8, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结表头
    ws.freeze_panes = f"A{header_row + 1}"

    # 打印设置
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"统计汇总表_{project['name']}_{timestamp}.xlsx"
    filepath = os.path.join(EXPORT_FOLDER, filename)
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    wb.save(filepath)

    return filepath


def export_individual(project_id, point_ids=None):
    """
    为每个调查点生成独立的调查记录表

    Returns: ZIP文件路径
    """
    project = get_project(project_id)
    if not project:
        raise ValueError("项目不存在")

    # 获取模板字段
    template_id = project.get("template_id")
    if template_id:
        from backend.data_manager import get_template
        template = get_template(template_id)
    else:
        template = get_default_template()

    if not template:
        raise ValueError("未找到调查字段模板")

    fields = template["fields"]

    # 获取调查点
    result = get_survey_points(project_id, per_page=10000)
    points = result["items"]
    if point_ids:
        pid_set = set(point_ids)
        points = [p for p in points if p["id"] in pid_set]

    if not points:
        raise ValueError("没有符合条件的调查点")

    # 状态映射
    status_map = {
        "pending": "待调查", "in_progress": "进行中",
        "surveyed": "已完成", "skipped": "已跳过"
    }

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    record_dir = os.path.join(EXPORT_FOLDER, f"原始记录表_{project['name']}_{timestamp}")
    os.makedirs(record_dir, exist_ok=True)

    for point in points:
        detail = get_point_detail(point["id"]) or {}
        records = {r["field_key"]: r["field_value"] for r in detail.get("records", [])}
        photos = detail.get("photos", [])

        wb = Workbook()
        ws = wb.active
        ws.title = "调查记录表"

        # === 表头 ===
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = "工程现场调查记录表"
        _apply_cell_style(title_cell, font=TITLE_FONT,
                          alignment=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[1].height = 35

        # === 基本信息区 ===
        ws.merge_cells("A3:D3")
        info_label = ws["A3"]
        info_label.value = "基本信息"
        _apply_cell_style(info_label, font=SUBTITLE_FONT)

        basic_fields = [
            ("编号", point["point_number"]),
            ("工程名称", project["name"]),
            ("调查日期", records.get("survey_date", point.get("surveyed_at", "")[:10] if point.get("surveyed_at") else "")),
            ("调查状态", status_map.get(point["status"], point["status"])),
        ]

        row = 4
        for label, value in basic_fields:
            cell_l = ws.cell(row=row, column=1, value=label)
            _apply_cell_style(cell_l, font=BOLD_FONT, fill=LABEL_FILL, alignment=CENTER_ALIGN, border=THIN_BORDER)

            cell_v = ws.cell(row=row, column=2, value=value)
            _apply_cell_style(cell_v, font=NORMAL_FONT, alignment=CENTER_ALIGN, border=THIN_BORDER)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

            ws.row_dimensions[row].height = 25
            row += 1

        # GPS 信息
        gps_label = ws.cell(row=row, column=1, value="GPS坐标")
        _apply_cell_style(gps_label, BOLD_FONT, LABEL_FILL, CENTER_ALIGN, THIN_BORDER)
        gps_val = ws.cell(
            row=row, column=2,
            value=f"经度: {point['longitude']:.6f}, 纬度: {point['latitude']:.6f}, 高程: {point['altitude']:.1f}m"
            if point["longitude"] else "未设置"
        )
        _apply_cell_style(gps_val, NORMAL_FONT, alignment=LEFT_ALIGN, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 25
        row += 2

        # === 调查内容区 ===
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        survey_label = ws.cell(row=row, column=1, value="调查内容")
        _apply_cell_style(survey_label, SUBTITLE_FONT)
        row += 1

        for field in fields:
            cell_l = ws.cell(row=row, column=1, value=field["field_label"])
            _apply_cell_style(cell_l, BOLD_FONT, LABEL_FILL, CENTER_ALIGN, THIN_BORDER)

            cell_v = ws.cell(row=row, column=2, value=records.get(field["field_key"], ""))
            _apply_cell_style(cell_v, NORMAL_FONT, alignment=LEFT_ALIGN, border=THIN_BORDER)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

            # 多行文本字段给更高行高
            if field["field_type"] == "multiline":
                ws.row_dimensions[row].height = 50
            else:
                ws.row_dimensions[row].height = 25

            row += 1

        row += 1

        # === 照片区 ===
        if photos:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            photo_label = ws.cell(row=row, column=1, value="现场照片")
            _apply_cell_style(photo_label, SUBTITLE_FONT)
            row += 1

            photo_col = 0
            photo_start_row = row
            for pi, photo in enumerate(photos):
                col = 1 + (photo_col % 2) * 3  # 每行2张，交替在A列和D列
                current_row = row

                # 照片说明
                caption = photo.get("caption", "") or f"照片{pi+1}"
                cap_cell = ws.cell(row=current_row, column=col, value=caption)
                _apply_cell_style(cap_cell, NORMAL_FONT, alignment=CENTER_ALIGN)
                ws.merge_cells(start_row=current_row, start_column=col,
                              end_row=current_row, end_column=col+2)
                current_row += 1

                # 尝试插入照片缩略图
                photo_path = None
                if photo.get("thumbnail_path"):
                    p = os.path.normpath(os.path.join(BASE_DIR, str(photo["thumbnail_path"])))
                    if os.path.isfile(p):
                        photo_path = p
                if not photo_path and photo.get("storage_path"):
                    p = os.path.normpath(os.path.join(BASE_DIR, str(photo["storage_path"])))
                    if os.path.isfile(p):
                        photo_path = p

                if photo_path:
                    img = _safe_xl_image(photo_path)
                    if img:
                        # 限制图片大小
                        img.width = min(img.width, 200)
                        img.height = min(img.height, 150)
                        anchor_cell = f"{get_column_letter(col)}{current_row}"
                        ws.add_image(img, anchor_cell)
                        current_row += 12  # 给图片留空间
                    else:
                        ws.cell(row=current_row, column=col,
                                value="[照片加载失败]")
                        current_row += 1
                else:
                    ws.cell(row=current_row, column=col,
                            value="[照片文件缺失]")
                    current_row += 1

                if photo_col % 2 == 1:
                    row = current_row  # 换行
                photo_col += 1

            # 奇数张照片时补一行
            if photo_col % 2 == 1:
                row = photo_start_row + max(
                    14 if photos else 0, 2 * len(photos)
                )

        # === 页脚 ===
        row += 2
        footer_font = Font(name="微软雅黑", size=9, color="888888")
        footer_cell = ws.cell(row=row, column=1,
                              value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
                                    f"奥维野外调查插件 v1.0")
        _apply_cell_style(footer_cell, footer_font)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

        # 设置列宽
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22

        # 打印设置
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # 保存
        safe_name = point["point_number"].replace("/", "_").replace("\\", "_").replace(":", "_")
        filepath = os.path.join(record_dir, f"{safe_name}.xlsx")
        wb.save(filepath)

    # 打包ZIP
    zip_filename = f"调查记录_{project['name']}_{timestamp}.zip"
    zip_path = os.path.join(EXPORT_FOLDER, zip_filename)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(record_dir):
            for file in files:
                full_path = os.path.join(root, file)
                arcname = os.path.relpath(full_path, EXPORT_FOLDER)
                zf.write(full_path, arcname)

    return zip_path


def export_all(project_id, point_ids=None):
    """
    导出所有表格（汇总表 + 单点记录表），打包为ZIP

    Returns: ZIP文件路径
    """
    project = get_project(project_id)
    if not project:
        raise ValueError("项目不存在")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 生成汇总表
    summary_path = export_summary(project_id, point_ids)

    # 生成单点记录表ZIP
    individual_zip = export_individual(project_id, point_ids)

    # 合并打包
    final_zip = os.path.join(
        EXPORT_FOLDER, f"完整导出_{project['name']}_{timestamp}.zip"
    )

    with zipfile.ZipFile(final_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        # 添加汇总表
        zf.write(summary_path, f"统计汇总表.xlsx")

        # 添加单点记录
        if os.path.exists(individual_zip):
            with zipfile.ZipFile(individual_zip, "r") as inner:
                for name in inner.namelist():
                    data = inner.read(name)
                    zf.writestr(f"原始记录表/{name}", data)

    return final_zip
